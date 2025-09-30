#!/usr/bin/env python3
"""
Excel Template Generator for Belofte Enterprise System
This script generates a comprehensive Excel template for bulk importing:
- Customers
- Customer Opening Balances 
- Suppliers
- Supplier Items & Prices
- Container Items (Inventory)
"""

import pandas as pd
from datetime import datetime, timedelta
import os
from pathlib import Path

def create_excel_template():
    """Creates a comprehensive Excel template with multiple worksheets"""
    
    # Create output directory if it doesn't exist
    output_dir = Path("./templates")
    output_dir.mkdir(exist_ok=True)
    
    # Create Excel writer object
    filename = f"Belofte_Import_Template_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = output_dir / filename
    
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        
        # 1. CUSTOMERS SHEET
        customers_data = {
            'customerName': [
                'John Smith Trading', 
                'ABC Corporation', 
                'Global Imports Ltd', 
                'Local Business Co'
            ],
            'phone': [
                '+1234567890', 
                '+1987654321', 
                '+1122334455', 
                '+1555666777'
            ]
        }
        customers_df = pd.DataFrame(customers_data)
        customers_df.to_excel(writer, sheet_name='Customers', index=False)
        
        # 2. CUSTOMER OPENING BALANCES SHEET
        opening_balances_data = {
            'customerName': [
                'John Smith Trading', 
                'ABC Corporation', 
                'Global Imports Ltd'
            ],
            'openingBalance': [1500.00, 2300.50, 750.25],
            'notes': [
                'Previous credit balance', 
                'Outstanding from last month', 
                'Partial payment pending'
            ]
        }
        opening_balances_df = pd.DataFrame(opening_balances_data)
        opening_balances_df.to_excel(writer, sheet_name='Customer Opening Balances', index=False)
        
        # 3. SUPPLIERS SHEET
        suppliers_data = {
            'suppliername': [
                'China Electronics Supplier',
                'European Parts Ltd', 
                'Asian Manufacturing Co',
                'Tech Components Inc'
            ],
            'contact': [
                '+86123456789', 
                '+44987654321', 
                '+65111222333', 
                '+1444555666'
            ],
            'country': [
                'China', 
                'United Kingdom', 
                'Singapore', 
                'United States'
            ]
        }
        suppliers_df = pd.DataFrame(suppliers_data)
        suppliers_df.to_excel(writer, sheet_name='Suppliers', index=False)
        
        # 4. SUPPLIER ITEMS & PRICES SHEET
        supplier_items_data = {
            'suppliername': [
                'China Electronics Supplier',
                'China Electronics Supplier', 
                'European Parts Ltd',
                'European Parts Ltd',
                'Asian Manufacturing Co',
                'Tech Components Inc'
            ],
            'itemName': [
                'iPhone 14 Pro Max',
                'Samsung Galaxy S23', 
                'MacBook Air M2',
                'iPad Pro 12.9"',
                'Sony WH-1000XM4',
                'Dell XPS 13'
            ],
            'price': [1099.99, 999.99, 1199.99, 1099.99, 349.99, 899.99]
        }
        supplier_items_df = pd.DataFrame(supplier_items_data)
        supplier_items_df.to_excel(writer, sheet_name='Supplier Items & Prices', index=False)
        
        # 5. CONTAINER ITEMS (INVENTORY) SHEET
        container_items_data = {
            'containerNo': [
                'CONT2024001',
                'CONT2024001', 
                'CONT2024002',
                'CONT2024002',
                'CONT2024003',
                'CONT2024003'
            ],
            'suppliername': [
                'China Electronics Supplier',
                'China Electronics Supplier',
                'European Parts Ltd', 
                'European Parts Ltd',
                'Asian Manufacturing Co',
                'Tech Components Inc'
            ],
            'itemName': [
                'iPhone 14 Pro Max',
                'Samsung Galaxy S23',
                'MacBook Air M2', 
                'iPad Pro 12.9"',
                'Sony WH-1000XM4',
                'Dell XPS 13'
            ],
            'quantity': [50, 30, 25, 20, 100, 15],
            'unitPrice': [1050.00, 950.00, 1150.00, 1050.00, 320.00, 850.00],
            'arrivalDate': [
                (datetime.now() - timedelta(days=10)).strftime('%Y-%m-%d'),
                (datetime.now() - timedelta(days=10)).strftime('%Y-%m-%d'),
                (datetime.now() - timedelta(days=5)).strftime('%Y-%m-%d'),
                (datetime.now() - timedelta(days=5)).strftime('%Y-%m-%d'),
                (datetime.now() - timedelta(days=2)).strftime('%Y-%m-%d'),
                datetime.now().strftime('%Y-%m-%d')
            ],
            'year': [2024, 2024, 2024, 2024, 2024, 2024]
        }
        container_items_df = pd.DataFrame(container_items_data)
        container_items_df.to_excel(writer, sheet_name='Container Items', index=False)
        
        # 6. INSTRUCTIONS SHEET
        instructions_data = {
            'Sheet Name': [
                'Customers',
                'Customer Opening Balances',
                'Suppliers', 
                'Supplier Items & Prices',
                'Container Items'
            ],
            'Description': [
                'Customer information with name and phone number',
                'Opening credit balances for existing customers',
                'Supplier information with contact details and country',
                'Items supplied by each supplier with their prices',
                'Container inventory with quantities and arrival dates'
            ],
            'Required Fields': [
                'customerName, phone',
                'customerName, openingBalance',
                'suppliername, contact, country',
                'suppliername, itemName, price', 
                'containerNo, suppliername, itemName, quantity, unitPrice, arrivalDate, year'
            ],
            'Notes': [
                'Customer names must be unique per company',
                'customerName must match existing customer or one from Customers sheet',
                'Supplier names must be unique per company',
                'suppliername must match existing supplier or one from Suppliers sheet',
                'arrivalDate format: YYYY-MM-DD, suppliername must exist'
            ]
        }
        instructions_df = pd.DataFrame(instructions_data)
        instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
    
    # Add formatting and validation to the Excel file
    format_excel_template(filepath)
    
    print(f"✅ Excel template created successfully: {filepath}")
    return str(filepath)

def format_excel_template(filepath):
    """Add formatting, colors, and data validation to the Excel template"""
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    try:
        wb = load_workbook(filepath)
        
        # Define colors and styles
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Format each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format headers
            if ws.max_row > 0:
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_alignment
                    
                    # Auto-size columns
                    column_letter = ws.cell(row=1, column=col).column_letter
                    ws.column_dimensions[column_letter].width = max(len(str(cell.value)) + 2, 12)
        
        wb.save(filepath)
        
    except Exception as e:
        print(f"Warning: Could not format Excel file: {e}")

if __name__ == "__main__":
    create_excel_template()